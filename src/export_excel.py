#!/usr/bin/env python3

from datetime import datetime

import openpyxl
from openpyxl.styles import Font


class ExcelExporter:
    def __init__(self, imessage_data: list, file_path: str):
        self.imessage_data = imessage_data
        self.file_path = file_path

    def export(self) -> str:
        users = []
        messages = []
        dates = []
        services = []
        accounts = []
        is_from_me = []

        for data in self.imessage_data:
            users.append(data.from_caller_id)
            messages.append(data.content)
            dates.append(data.sent_on)
            services.append(data.service)
            accounts.append(data.to_caller_id)
            is_from_me.append(data.is_from_me)

        # Call openpyxl.Workbook() to create a new blank Excel workbook
        workbook = openpyxl.Workbook()

        # Activate a sheet
        sheet = workbook.active

        # Set a title
        sheet.title = "iMessages"

        # Set headline style
        bold16font = Font(size=16, bold=True)

        sheet["A1"] = "User ID"
        sheet["A1"].font = bold16font

        sheet["B1"] = "Message"
        sheet["B1"].font = bold16font

        sheet["C1"] = "Date"
        sheet["C1"].font = bold16font

        sheet["D1"] = "Service"
        sheet["D1"].font = bold16font

        sheet["E1"] = "Destination Caller ID"
        sheet["E1"].font = bold16font

        sheet["F1"] = "Is From Me"
        sheet["F1"].font = bold16font

        # Write users to 1st column
        users_row = 2
        for user in users:
            sheet.cell(row=users_row, column=1).value = user
            users_row += 1

        # Write messages to 2nd column
        messages_row = 2
        for message in messages:
            sheet.cell(row=messages_row, column=2).value = message
            messages_row += 1

        # Write date to 3rd column
        dates_row = 2
        for date in dates:
            sheet.cell(row=dates_row, column=3).value = date
            dates_row += 1

        # Write services to 4th column
        service_row = 2
        for service in services:
            sheet.cell(row=service_row, column=4).value = service
            service_row += 1

        # Write accounts to 5th column
        account_row = 2
        for account in accounts:
            sheet.cell(row=account_row, column=5).value = account
            account_row += 1

        # Write is_from_me to 6th column
        is_from_me_row = 2
        for from_me in is_from_me:
            sheet.cell(row=is_from_me_row, column=6).value = from_me
            is_from_me_row += 1

        # Save the workbook
        filename = (
            f'{self.file_path}iMessage-Data_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
        )
        workbook.save(filename)
        return filename
